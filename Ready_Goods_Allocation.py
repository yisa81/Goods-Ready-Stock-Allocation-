import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math

def balance_conant_ocean_qty(df):
    for idx, row in df.iterrows():
        try:
            conant_soh = row['conant soh'] or 0
            ocean_soh = row['ocean soh'] or 0
            conant_qty = row['conant qty'] or 0
            ocean_qty = row['ocean qty'] or 0
            monthly_max_avg = row['mthly max avg sales (a,b & c)'] or 1
            count = row['count'] or 1

            conant_msoh = (conant_soh + conant_qty) / (monthly_max_avg * 0.6)
            ocean_msoh = (ocean_soh + ocean_qty) / (monthly_max_avg * 0.4)

            tolerance = 0.05

            if count == 2:
                target_gap = 0.1
            else:
                target_gap = 0.0

            max_iterations = 50
            iterations = 0

            while abs((conant_msoh - ocean_msoh) - target_gap) > tolerance and iterations < max_iterations:
                move_qty = 5
                if (conant_msoh - ocean_msoh) > target_gap:
                    if df.at[idx, 'conant qty'] > move_qty:
                        df.at[idx, 'conant qty'] -= move_qty
                        df.at[idx, 'ocean qty'] += move_qty
                else:
                    if df.at[idx, 'ocean qty'] > move_qty:
                        df.at[idx, 'ocean qty'] -= move_qty
                        df.at[idx, 'conant qty'] += move_qty

                conant_msoh = (conant_soh + df.at[idx, 'conant qty']) / (monthly_max_avg * 0.6)
                ocean_msoh = (ocean_soh + df.at[idx, 'ocean qty']) / (monthly_max_avg * 0.4)
                iterations += 1

        except Exception:
            pass
    return df

st.title("🛒 Goods Ready Stock Allocation (Final Corrected Version)")

sales_file = st.file_uploader("Upload Sales Report (.xlsx)", type="xlsx")
goods_file = st.file_uploader("Upload Ready Goods (.xlsx)", type="xlsx")

if sales_file and goods_file:
    try:
        sales_xls = pd.ExcelFile(sales_file)

        if not {'Data', 'OOS'}.issubset(set(sales_xls.sheet_names)):
            st.error("Sales Report must contain 'Data' and 'OOS' sheets!")
            st.stop()

        data_df = pd.read_excel(sales_xls, sheet_name='Data', engine="openpyxl")
        oos_df = pd.read_excel(sales_xls, sheet_name='OOS', engine="openpyxl")
        goods_df = pd.read_excel(goods_file, engine="openpyxl")

        data_df.columns = data_df.columns.str.strip().str.lower().str.replace('.', '')
        oos_df.columns = oos_df.columns.str.strip().str.lower().str.replace('.', '')
        goods_df.columns = goods_df.columns.str.strip().str.lower().str.replace('.', '')

        goods_po_unique = goods_df[['po no']].drop_duplicates()
        matched_po_df = pd.merge(goods_po_unique, oos_df, how='inner', on='po no')

        st.subheader("🔍 Matching PO No. and their SKUs from OOS Sheet:")
        st.dataframe(matched_po_df[['po no', 'simple sku']])

        merged_df = matched_po_df.merge(data_df, left_on='simple sku', right_on='sku', how='left')

        merged_df['actual outstanding balance'] = merged_df['actual outstanding balance'].fillna(0)
        merged_df['conant qty'] = 0
        merged_df['ocean qty'] = 0

        for idx, row in merged_df.iterrows():
            supplier = row['supplier_x'] or ''
            outstanding_balance = row['actual outstanding balance'] or 0
            count = row['count'] or 1
            conant_soh = row['conant soh'] or 0
            ocean_soh = row['ocean soh'] or 0
            monthly_max = row['mthly max avg sales (a,b & c)'] or 1

            if 'ocean' in supplier.lower():
                merged_df.at[idx, 'ocean qty'] = outstanding_balance
            else:
                conant_msoh = (conant_soh) / (monthly_max * 0.6)
                ocean_msoh = (ocean_soh) / (monthly_max * 0.4)

                if outstanding_balance < 50:
                    merged_df.at[idx, 'conant qty'] = outstanding_balance
                    merged_df.at[idx, 'ocean qty'] = 0
                else:
                    best_conant_qty = 0
                    best_score = None

                    for conant_qty_candidate in range(50, outstanding_balance + 1, 50):
                        ocean_qty_candidate = outstanding_balance - conant_qty_candidate
                        conant_msoh_candidate = (conant_soh + conant_qty_candidate) / (monthly_max * 0.6)
                        ocean_msoh_candidate = (ocean_soh + ocean_qty_candidate) / (monthly_max * 0.4)

                        if count == 2:
                            if conant_msoh_candidate > ocean_msoh_candidate:
                                score = conant_msoh_candidate - ocean_msoh_candidate
                                if best_score is None or score > best_score:
                                    best_score = score
                                    best_conant_qty = conant_qty_candidate
                        else:
                            difference = abs(conant_msoh_candidate - ocean_msoh_candidate)
                            if best_score is None or difference < best_score:
                                best_score = difference
                                best_conant_qty = conant_qty_candidate

                    merged_df.at[idx, 'conant qty'] = best_conant_qty
                    merged_df.at[idx, 'ocean qty'] = outstanding_balance - best_conant_qty

        merged_df = balance_conant_ocean_qty(merged_df)

        merged_df['conant msoh'] = ""
        merged_df['ocean msoh'] = ""

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            merged_df.to_excel(writer, index=False, sheet_name="Data")

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Data"]

        headers = list(merged_df.columns)
        row_count = ws.max_row

        col_conant_soh = get_column_letter(headers.index("conant soh") + 1)
        col_conant_qty = get_column_letter(headers.index("conant qty") + 1)
        col_mthly_max = get_column_letter(headers.index("mthly max avg sales (a,b & c)") + 1)
        col_conant_msoh = get_column_letter(headers.index("conant msoh") + 1)
        col_ocean_soh = get_column_letter(headers.index("ocean soh") + 1)
        col_ocean_qty = get_column_letter(headers.index("ocean qty") + 1)
        col_ocean_msoh = get_column_letter(headers.index("ocean msoh") + 1)
        ready_to_ship_col = headers.index("ready to ship") + 1 if "ready to ship" in headers else None
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in range(2, row_count + 1):
            ws[f"{col_conant_msoh}{row}"].value = f"=ROUND(({col_conant_soh}{row}+{col_conant_qty}{row})/({col_mthly_max}{row}*0.6),2)"
            ws[f"{col_ocean_msoh}{row}"].value = f"=ROUND(({col_ocean_soh}{row}+{col_ocean_qty}{row})/({col_mthly_max}{row}*0.4),2)"
            if ready_to_ship_col:
                ws.cell(row=row, column=ready_to_ship_col).fill = yellow_fill

        columns_to_hide = [2, (4, 6), 9, (11,12), (15, 38), (40, 45), 49, (51, 69), (71, 74)] 
        for col in columns_to_hide:
            if isinstance(col, tuple):
                col_start, col_end = col
                for c in range(col_start, col_end + 1):
                    ws.column_dimensions[get_column_letter(c)].hidden = True
            else:
                ws.column_dimensions[get_column_letter(col)].hidden = True

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.success("✅ File processed, balanced, optimized, and finalized with correct MSOH balancing!")
        st.download_button(
            label="📥 Download Final Excel",
            data=final_output,
            file_name="Merged_Goods_Allocation_Final_Balanced.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Something went wrong: {e}")
